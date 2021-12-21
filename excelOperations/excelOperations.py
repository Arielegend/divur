from openpyxl import load_workbook

helperAddfilePath = 'HelperAdd.xlsx'


# portsNames, portCodes, regions, states, carriers, dv20Ashdod, dv40Ashdod, hq40Ashdod, dv20Haifa, dv40Haifa, hq40Haifa, validFrom, validUntil
def Load_New_Fcl_Rates_From_Excel(xl):
    portsNames = xl['PortName'].values
    portCodes = xl['PortCode'].values
    regions = xl['Region'].values
    states = xl['State'].values
    carriers = xl['Carrier'].values

    dv20Ashdod = xl['Dv20Ashdod'].values
    dv40Ashdod = xl['Dv40Ashdod'].values
    hq40Ashdod = xl['Hq40Ashdod'].values

    dv20Haifa = xl['Dv20Haifa'].values
    dv40Haifa = xl['Dv40Haifa'].values
    hq40Haifa = xl['Hq40Haifa'].values

    validFrom = xl['ValidFrom'].values
    validUntil = xl['ValidUntil'].values

    allRows = []
    for (portName, portCode, region, state, carrier, dv20_Ashdod, dv40_Ashdod, hq40_Ashdod, dv20_Haifa, dv40_Haifa,
         hq40_Haifa, validStart, validEnd) \
            in zip(portsNames, portCodes, regions, states, carriers, dv20Ashdod, dv40Ashdod, hq40Ashdod, dv20Haifa,
                   dv40Haifa, hq40Haifa, validFrom, validUntil):
        allRows.append((portName, portCode, region, state, carrier, dv20_Ashdod, dv40_Ashdod, hq40_Ashdod, dv20_Haifa,
                        dv40_Haifa, hq40_Haifa, validStart, validEnd))

    return allRows


# FirstName	LastName	Email	Phone	DeltaFcl	DeltaLcl	DeltaAir	RelevantPorts	Info
def Load_New_Clients_From_Excel(xl):
    firstNames = xl['FirstName'].values
    lastName = xl['LastName'].values
    emails = xl['Email'].values
    phones = xl['Phone'].values

    deltasFcl = xl['DeltaFcl'].values
    deltasLcl = xl['DeltaLcl'].values
    deltasAir = xl['DeltaAir'].values

    relevantPorts = xl['RelevantPorts'].values
    infos = xl['Info'].values

    allRows = []
    for (firstName, lastName, email, phone, deltaFcl, deltaLcl, deltaAir, relevantPortsHelper, info) \
            in zip(firstNames, lastName, emails, phones, deltasFcl, deltasLcl, deltasAir, relevantPorts, infos):
        helper = (firstName, lastName, email, phone, deltaFcl, deltaLcl, deltaAir, relevantPortsHelper, info)
        allRows.append(helper)

    return allRows


def Clear_Sheet(sheetName: str):
    book = load_workbook(helperAddfilePath)
    sheet = book[sheetName]

    Delete_Sheet_But_Headers(sheet)
    book.save(helperAddfilePath)


def Delete_Sheet_But_Headers(sheet):
    while sheet.max_row > 1:
        sheet.delete_rows(2)
    return


def Load_Ports_From_Excel(xl):
    portNames = xl['PortName'].values
    portCodes = xl['PortCode'].values
    portsStates = xl['PortState'].values
    portsAliases = xl['PortAliases'].values

    allRows = []
    for (portName, portCode, portsState, portAliases) in zip(portNames, portCodes, portsStates, portsAliases):
        allRows.append((portName, portCode, portsState, portAliases))

    return allRows
